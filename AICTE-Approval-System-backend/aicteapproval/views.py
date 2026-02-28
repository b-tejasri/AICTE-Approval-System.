from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.shortcuts import render
from django.http import HttpResponse
from datetime import datetime
from django.conf import settings
import os
from django.shortcuts import render, redirect
from .models import *
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import generics, status

from django.conf import settings
from django.db.models import Q  #  THIS FIXES YOUR ERROR

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAdminUser
from rest_framework import status
from rest_framework.generics import ListAPIView
from rest_framework.permissions import AllowAny
from rest_framework.generics import RetrieveAPIView
import uuid
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response

import json
import hmac
import hashlib
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
import re
import fitz  # PyMuPDF
from groq import Groq
from django.views import View

# ── Groq client ────────────────────────────────────────  uncomment this 
groq_client = Groq(api_key="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")


# ── helpers ────────────────────────────────────────────


from django.utils.decorators import method_decorator

import io
from datetime import datetime
from django.http import HttpResponse
from django.views import View

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ── colour constants ──────────────────────────────────────────────────────────
NAVY_HEX   = '0A2240'
SAFF_HEX   = 'FF6B00'
GREEN_HEX  = '138808'
AMBER_HEX  = 'E8920A'
RED_HEX    = 'D92B3A'
GRAY_HEX   = 'F4F6F9'
LGRAY_HEX  = 'E8ECF2'
WHITE_HEX  = 'FFFFFF'
GOLD_HEX   = 'C8A84B'

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _font(bold=False, color=None, size=10, name='Arial'):
    return Font(name=name, bold=bold, color=color or '000000', size=size)

def _border():
    thin = Side(style='thin', color='D0D7E3')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def _apply_header_row(ws, row_num, values, col_widths=None):
    """Write a styled navy header row."""
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=ci, value=val)
        c.font      = _font(bold=True, color=WHITE_HEX, size=9)
        c.fill      = _fill(NAVY_HEX)
        c.alignment = _center()
        c.border    = _border()
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

def _apply_data_row(ws, row_num, values, shade=False):
    """Write a styled data row with alternating shading."""
    bg = GRAY_HEX if shade else WHITE_HEX
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=ci, value=val)
        c.font      = _font(size=9)
        c.fill      = _fill(bg)
        c.alignment = _left()
        c.border    = _border()

def _section_title(ws, row_num, title, ncols):
    """Write an orange section title spanning ncols."""
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=ncols)
    c = ws.cell(row=row_num, column=1, value=title)
    c.font      = _font(bold=True, color=WHITE_HEX, size=10)
    c.fill      = _fill(SAFF_HEX)
    c.alignment = _center()
    c.border    = _border()
    ws.row_dimensions[row_num].height = 18


class DownloadInstitutionExcelView(View):
    """
    GET /api/.../download-excel/?institution_id=<id>
    Returns a fully formatted .xlsx with all institution data split across sheets.
    """

    def get(self, request):
        inst_id = request.GET.get('institution_id')
        if not inst_id:
            return HttpResponse('institution_id required', status=400)

        try:
            inst = Institution.objects.get(pk=inst_id)
        except Institution.DoesNotExist:
            return HttpResponse('Institution not found', status=404)

        try:
            data = inst.data
        except InstitutionData.DoesNotExist:
            data = None

        latest_risk = AIRiskAnalysis.objects.filter(institution=inst).order_by('-analyzed_at').first()
        disclosures  = DisclosureSection.objects.filter(institution=inst).order_by('section_type', '-uploaded_at')
        all_risk     = AIRiskAnalysis.objects.filter(institution=inst).order_by('-analyzed_at')

        wb = Workbook()
        wb.remove(wb.active)   # remove default sheet

        self._sheet_overview(wb, inst, data, latest_risk)
        self._sheet_faculty(wb, data)
        self._sheet_labs(wb, data)
        self._sheet_infrastructure(wb, data)
        self._sheet_students(wb, data)
        self._sheet_financials(wb, data)
        self._sheet_accreditation(wb, data)
        self._sheet_risk_analysis(wb, latest_risk, all_risk)
        self._sheet_disclosures_log(wb, disclosures)

        # ── stream response ──────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = inst.institution_name.replace(' ', '_')[:40]
        filename  = f"AICTE_{safe_name}_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Access-Control-Allow-Origin'] = '*'
        return response

    # ── SHEET 1: OVERVIEW ────────────────────────────────────────────────────
    def _sheet_overview(self, wb, inst, data, risk):
        ws = wb.create_sheet('Overview')
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 42

        # Big title block
        ws.merge_cells('A1:B1')
        c = ws['A1']
        c.value     = 'AICTE Compliance Report'
        c.font      = _font(bold=True, color=WHITE_HEX, size=16, name='Arial')
        c.fill      = _fill(NAVY_HEX)
        c.alignment = _center()
        ws.row_dimensions[1].height = 32

        ws.merge_cells('A2:B2')
        c = ws['A2']
        c.value     = inst.institution_name
        c.font      = _font(bold=True, color=WHITE_HEX, size=13)
        c.fill      = _fill(SAFF_HEX)
        c.alignment = _center()
        ws.row_dimensions[2].height = 24

        ws.merge_cells('A3:B3')
        c = ws['A3']
        c.value     = f'Generated: {datetime.now().strftime("%d %B %Y, %I:%M %p")} | Academic Year: 2024-25'
        c.font      = _font(color='555555', size=8)
        c.fill      = _fill(LGRAY_HEX)
        c.alignment = _center()
        ws.row_dimensions[3].height = 14

        row = 5
        _section_title(ws, row, 'INSTITUTION PROFILE', 2); row += 1

        profile_rows = [
            ('Institution Name',    inst.institution_name),
            ('AICTE ID',            inst.aicte_id or '—'),
            ('Type',                inst.inst_type),
            ('Category',            inst.category),
            ('Year Established',    str(inst.year_established)),
            ('Affiliated University', inst.affiliated_univ or '—'),
            ('State',               inst.state),
            ('District',            inst.district),
            ('Principal',           inst.principal_name or '—'),
            ('Email',               inst.email),
            ('Mobile',              inst.mobile or '—'),
        ]
        for i, (lbl, val) in enumerate(profile_rows):
            ws.cell(row=row, column=1, value=lbl).font = _font(bold=True, size=9)
            ws.cell(row=row, column=1).fill  = _fill(GRAY_HEX)
            ws.cell(row=row, column=1).border = _border()
            ws.cell(row=row, column=2, value=val).font = _font(size=9)
            ws.cell(row=row, column=2).border = _border()
            row += 1

        row += 1
        _section_title(ws, row, 'KEY STATISTICS', 2); row += 1
        _apply_header_row(ws, row, ['Metric', 'Value']); row += 1

        stats = []
        if data:
            stats = [
                ('Total Faculty',       data.total_faculty),
                ('Required Faculty',    data.required_faculty),
                ('Faculty Shortage',    max(0, data.required_faculty - data.total_faculty)),
                ('PhD Holders',         data.faculty_phd_count),
                ('Total Students',      data.total_students),
                ('UG Students',         data.ug_students),
                ('PG Students',         data.pg_students),
                ('Total Laboratories',  data.total_labs),
                ('Total Classrooms',    data.total_classrooms),
                ('Library Books',       data.library_books),
                ('Total Computers',     data.computer_count),
                ('Total Area (sqft)',   data.total_area_sqft),
                ('Hostel Capacity',     data.hostel_capacity),
                ('Annual Budget (Lakhs)', data.annual_budget),
                ('NAAC Grade',          data.naac_grade or '—'),
                ('ISO Certified',       'Yes' if data.iso_certified else 'No'),
            ]
        for i, (lbl, val) in enumerate(stats):
            _apply_data_row(ws, row, [lbl, val], shade=(i % 2 == 0)); row += 1

        row += 1
        _section_title(ws, row, 'AI RISK SUMMARY', 2); row += 1
        _apply_header_row(ws, row, ['Risk Parameter', 'Value']); row += 1
        if risk:
            risk_color = RED_HEX if risk.risk_score >= 70 else (AMBER_HEX if risk.risk_score >= 40 else GREEN_HEX)
            risk_rows = [
                ('Risk Score',        f'{risk.risk_score} / 100'),
                ('Risk Level',        risk.risk_level),
                ('Compliance %',      f'{round(risk.compliance_pct, 1)}%'),
                ('Faculty Shortage',  'Yes' if risk.faculty_shortage else 'No'),
                ('Infra Deficit',     'Yes' if risk.infra_deficit else 'No'),
                ('Expired Certs',     'Yes' if risk.expired_certs else 'No'),
                ('Faculty Ratio',     f'1:{round(risk.faculty_ratio, 1)}' if risk.faculty_ratio else 'N/A'),
                ('Analyzed On',       risk.analyzed_at.strftime('%d %b %Y')),
            ]
            for i, (lbl, val) in enumerate(risk_rows):
                _apply_data_row(ws, row, [lbl, val], shade=(i % 2 == 0))
                # colour the score cell
                if lbl == 'Risk Score':
                    ws.cell(row=row, column=2).font = _font(bold=True, color=risk_color, size=10)
                row += 1
        else:
            ws.cell(row=row, column=1, value='No risk analysis available yet.').font = _font(color='888888')
            row += 1

    # ── SHEET 2: FACULTY ─────────────────────────────────────────────────────
    def _sheet_faculty(self, wb, data):
        ws = wb.create_sheet('Faculty Details')
        ws.sheet_view.showGridLines = False

        row = 1
        _section_title(ws, row, 'FACULTY DETAILS — MANDATORY DISCLOSURE 2024-25', 6); row += 2

        _section_title(ws, row, 'Faculty Summary', 6); row += 1
        _apply_header_row(ws, row, ['Parameter', 'Value'], col_widths=[28, 18, 20, 20, 18, 18]); row += 1
        summary = [
            ('Total Faculty',    data.total_faculty if data else 0),
            ('Required Faculty', data.required_faculty if data else 0),
            ('Shortage',         f'=B{row}-B{row-1}' if data else 0),
            ('PhD Holders',      data.faculty_phd_count if data else 0),
        ]
        for i, (lbl, val) in enumerate(summary):
            _apply_data_row(ws, row, [lbl, val], shade=(i % 2 == 0)); row += 1

        row += 1
        _section_title(ws, row, 'Faculty List (from PDF)', 6); row += 1
        _apply_header_row(ws, row, ['S.No', 'Name', 'Department', 'Qualification', 'Experience (Yrs)', 'Specialization']); row += 1

        faculty_list = (data.faculty_details if data else []) or []
        if faculty_list:
            for i, f in enumerate(faculty_list):
                _apply_data_row(ws, row, [
                    i + 1,
                    f.get('name', '—'),
                    f.get('dept', '—'),
                    f.get('qualification', '—'),
                    f.get('experience_years', 0),
                    f.get('specialization', '—'),
                ], shade=(i % 2 == 0))
                row += 1
        else:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1, value='No faculty data uploaded yet. Upload Faculty PDF to populate.').font = _font(color='888888', size=9)
            row += 1

    # ── SHEET 3: LABS ────────────────────────────────────────────────────────
    def _sheet_labs(self, wb, data):
        ws = wb.create_sheet('Laboratory Details')
        ws.sheet_view.showGridLines = False

        row = 1
        _section_title(ws, row, 'LABORATORY DETAILS — MANDATORY DISCLOSURE 2024-25', 5); row += 2

        _section_title(ws, row, 'Lab Summary', 5); row += 1
        _apply_header_row(ws, row, ['Parameter', 'Value'], col_widths=[28, 20, 24, 20, 18]); row += 1
        _apply_data_row(ws, row, ['Total Labs', data.total_labs if data else 0]); row += 1

        row += 1
        _section_title(ws, row, 'Lab List (from PDF)', 5); row += 1
        _apply_header_row(ws, row, ['S.No', 'Lab Name', 'Department', 'Area (sqft)', 'Equipment Count']); row += 1

        lab_list = (data.lab_details if data else []) or []
        if lab_list:
            for i, lab in enumerate(lab_list):
                _apply_data_row(ws, row, [
                    i + 1,
                    lab.get('name', '—'),
                    lab.get('dept', '—'),
                    lab.get('area_sqft', 0),
                    lab.get('equipment_count', 0),
                ], shade=(i % 2 == 0))
                row += 1

            # Totals row
            total_row = row
            ws.cell(row=total_row, column=1, value='TOTAL').font = _font(bold=True, size=9)
            ws.cell(row=total_row, column=1).fill = _fill(NAVY_HEX)
            ws.cell(row=total_row, column=1).font = _font(bold=True, color=WHITE_HEX, size=9)
            ws.cell(row=total_row, column=4, value=f'=SUM(D{row - len(lab_list)}:D{row - 1})').font = _font(bold=True, size=9)
            ws.cell(row=total_row, column=5, value=f'=SUM(E{row - len(lab_list)}:E{row - 1})').font = _font(bold=True, size=9)
            for ci in range(1, 6):
                ws.cell(row=total_row, column=ci).fill   = _fill(LGRAY_HEX)
                ws.cell(row=total_row, column=ci).border = _border()
        else:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.cell(row=row, column=1, value='No lab data uploaded yet. Upload Lab PDF to populate.').font = _font(color='888888', size=9)

    # ── SHEET 4: INFRASTRUCTURE ──────────────────────────────────────────────
    def _sheet_infrastructure(self, wb, data):
        ws = wb.create_sheet('Infrastructure')
        ws.sheet_view.showGridLines = False

        row = 1
        _section_title(ws, row, 'INFRASTRUCTURE DETAILS — MANDATORY DISCLOSURE 2024-25', 4); row += 2

        _apply_header_row(ws, row, ['Parameter', 'Available', 'AICTE Minimum', 'Status'],
                          col_widths=[32, 22, 22, 20]); row += 1

        if data:
            infra_rows = [
                ('Total Classrooms',           data.total_classrooms, 48,
                 'Compliant' if data.total_classrooms >= 48 else 'Below Norm'),
                ('Library Books (Volumes)',     data.library_books, 10000,
                 'Compliant' if data.library_books >= 10000 else 'Below Norm'),
                ('Total Computers',             data.computer_count, 300,
                 'Compliant' if data.computer_count >= 300 else 'Below Norm'),
                ('Total Area (sqft)',           data.total_area_sqft, 50000,
                 'Compliant' if data.total_area_sqft >= 50000 else 'Below Norm'),
                ('Hostel Capacity',             data.hostel_capacity, 0,
                 'Available' if data.hostel_capacity > 0 else 'Not Available'),
            ]
            for i, (lbl, avail, req, status) in enumerate(infra_rows):
                _apply_data_row(ws, row, [lbl, avail, req if req else '—', status], shade=(i % 2 == 0))
                status_cell = ws.cell(row=row, column=4)
                status_cell.font = _font(bold=True,
                    color=GREEN_HEX if 'Compliant' in status or 'Available' in status else RED_HEX,
                    size=9)
                row += 1
        else:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value='No infrastructure data yet.').font = _font(color='888888', size=9)

    # ── SHEET 5: STUDENTS ────────────────────────────────────────────────────
    def _sheet_students(self, wb, data):
        ws = wb.create_sheet('Student Details')
        ws.sheet_view.showGridLines = False

        row = 1
        _section_title(ws, row, 'STUDENT DETAILS — MANDATORY DISCLOSURE 2024-25', 3); row += 2

        _section_title(ws, row, 'Enrollment Summary', 3); row += 1
        _apply_header_row(ws, row, ['Parameter', 'Count', 'Remarks'], col_widths=[30, 18, 36]); row += 1

        if data:
            enroll = [
                ('Total Students',  data.total_students, ''),
                ('UG Students',     data.ug_students,    'B.Tech / B.E.'),
                ('PG Students',     data.pg_students,    'M.Tech / MBA / MCA'),
            ]
            for i, (lbl, val, rem) in enumerate(enroll):
                _apply_data_row(ws, row, [lbl, val, rem], shade=(i % 2 == 0)); row += 1

            row += 1
            _section_title(ws, row, 'Programs Offered', 3); row += 1
            _apply_header_row(ws, row, ['S.No', 'Program Name', 'Level']); row += 1
            programs = data.programs_offered or []
            for i, prog in enumerate(programs):
                level = 'PG' if any(x in prog for x in ['M.Tech', 'MBA', 'MCA', 'M.E.', 'M.Sc']) else 'UG'
                _apply_data_row(ws, row, [i + 1, prog, level], shade=(i % 2 == 0)); row += 1
        else:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.cell(row=row, column=1, value='No student data yet.').font = _font(color='888888', size=9)

    # ── SHEET 6: FINANCIALS ──────────────────────────────────────────────────
    def _sheet_financials(self, wb, data):
        ws = wb.create_sheet('Financial Details')
        ws.sheet_view.showGridLines = False

        row = 1
        _section_title(ws, row, 'FINANCIAL DETAILS — MANDATORY DISCLOSURE 2024-25', 3); row += 2

        _apply_header_row(ws, row, ['Parameter', 'Amount (Rs.)', 'Remarks'], col_widths=[30, 24, 36]); row += 1

        if data:
            fee = data.fee_structure or {}
            fin_rows = [
                ('Annual Budget',       f'Rs. {data.annual_budget} Lakhs', 'Total operational budget'),
                ('UG Tuition Fee',      f'Rs. {fee.get("ug_fee", 0):,}',   'Per annum'),
                ('PG Tuition Fee',      f'Rs. {fee.get("pg_fee", 0):,}',   'Per annum'),
            ]
            for i, (lbl, val, rem) in enumerate(fin_rows):
                _apply_data_row(ws, row, [lbl, val, rem], shade=(i % 2 == 0)); row += 1
        else:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.cell(row=row, column=1, value='No financial data yet.').font = _font(color='888888', size=9)

    # ── SHEET 7: ACCREDITATION ───────────────────────────────────────────────
    def _sheet_accreditation(self, wb, data):
        ws = wb.create_sheet('Accreditation')
        ws.sheet_view.showGridLines = False

        row = 1
        _section_title(ws, row, 'ACCREDITATION DETAILS — MANDATORY DISCLOSURE 2024-25', 3); row += 2

        _apply_header_row(ws, row, ['Parameter', 'Details', 'Status'], col_widths=[28, 36, 20]); row += 1

        if data:
            acc_rows = [
                ('NAAC Grade',      data.naac_grade or '—',     'Active' if data.naac_grade else 'Not Available'),
                ('NBA Programs',    data.nba_programs or '—',   'Active' if data.nba_programs else 'Not Available'),
                ('ISO 9001:2015',   'Certified' if data.iso_certified else 'Not Certified',
                                    'Valid' if data.iso_certified else 'Not Available'),
            ]
            for i, (lbl, val, status) in enumerate(acc_rows):
                _apply_data_row(ws, row, [lbl, val, status], shade=(i % 2 == 0))
                ws.cell(row=row, column=3).font = _font(
                    bold=True,
                    color=GREEN_HEX if status in ('Active', 'Valid') else AMBER_HEX,
                    size=9)
                row += 1
        else:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.cell(row=row, column=1, value='No accreditation data yet.').font = _font(color='888888', size=9)

    # ── SHEET 8: AI RISK ANALYSIS ────────────────────────────────────────────
    def _sheet_risk_analysis(self, wb, latest_risk, all_risk):
        ws = wb.create_sheet('AI Risk Analysis')
        ws.sheet_view.showGridLines = False

        row = 1
        _section_title(ws, row, 'AI RISK & COMPLIANCE ANALYSIS', 4); row += 2

        if not latest_risk:
            ws.cell(row=row, column=1, value='No risk analysis yet. Upload disclosure PDFs first.').font = _font(color='888888')
            return

        # Latest Risk Summary
        _section_title(ws, row, 'Latest Risk Summary', 4); row += 1
        _apply_header_row(ws, row, ['Parameter', 'Value', 'Flag', 'Remarks'],
                          col_widths=[28, 20, 20, 42]); row += 1

        risk_color = RED_HEX if latest_risk.risk_score >= 70 else (AMBER_HEX if latest_risk.risk_score >= 40 else GREEN_HEX)
        summary_rows = [
            ('Risk Score',        f'{latest_risk.risk_score} / 100',
             '🔴 HIGH' if latest_risk.risk_score >= 70 else ('🟡 MEDIUM' if latest_risk.risk_score >= 40 else '🟢 LOW'),
             'Higher score = Higher risk'),
            ('Risk Level',        latest_risk.risk_level,   '—',  ''),
            ('Compliance %',      f'{round(latest_risk.compliance_pct, 1)}%', '—', ''),
            ('Faculty Shortage',  'YES' if latest_risk.faculty_shortage else 'NO',
             '⚠️ Action Needed' if latest_risk.faculty_shortage else '✅ OK', ''),
            ('Infra Deficit',     'YES' if latest_risk.infra_deficit else 'NO',
             '⚠️ Action Needed' if latest_risk.infra_deficit else '✅ OK', ''),
            ('Expired Certs',     'YES' if latest_risk.expired_certs else 'NO',
             '⚠️ Action Needed' if latest_risk.expired_certs else '✅ OK', ''),
            ('Faculty Ratio',     f'1:{round(latest_risk.faculty_ratio, 1)}' if latest_risk.faculty_ratio else 'N/A',
             '⚠️ Above 1:15' if latest_risk.faculty_ratio > 15 else '✅ Within Norm', 'AICTE norm: 1:15'),
            ('Analyzed On',       latest_risk.analyzed_at.strftime('%d %b %Y, %I:%M %p'), '—', ''),
        ]
        for i, row_data in enumerate(summary_rows):
            _apply_data_row(ws, row, list(row_data), shade=(i % 2 == 0))
            if row_data[0] == 'Risk Score':
                ws.cell(row=row, column=2).font = _font(bold=True, color=risk_color, size=10)
            row += 1

        # Risk Factors
        row += 1
        _section_title(ws, row, 'Risk Factors Identified', 4); row += 1
        _apply_header_row(ws, row, ['S.No', 'Risk Factor', '', '']); row += 1
        for i, factor in enumerate(latest_risk.risk_factors or []):
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
             # Column 1 (S.No)
            _apply_data_row(ws, row, [i + 1], shade=(i % 2 == 0))

    # Only write to the top-left cell of merged region
            c = ws.cell(row=row, column=2, value=factor)
            c.font = _font(color=RED_HEX, size=9)
            c.alignment = _left()
            c.fill = _fill(GRAY_HEX if (i % 2 == 0) else WHITE_HEX)
            c.border = _border()

            row += 1
        row += 1
        _section_title(ws, row, 'Suggested Actions', 4); row += 1
        _apply_header_row(ws, row, ['S.No', 'Action Required', '', '']); row += 1
        for i, sug in enumerate(latest_risk.suggestions or []):
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
                          # S.No
            c1 = ws.cell(row=row, column=1, value=i + 1)
            c1.font = _font(size=9)
            c1.fill = _fill(GRAY_HEX if i % 2 == 0 else WHITE_HEX)
            c1.alignment = _left()
            c1.border = _border()

        # Only write to FIRST merged cell
            c2 = ws.cell(row=row, column=2, value=sug)
            c2.font = _font(color=GREEN_HEX, size=9)
            c2.fill = _fill(GRAY_HEX if i % 2 == 0 else WHITE_HEX)
            c2.alignment = _left()
            c2.border = _border()

            row += 1

        row += 1
        _section_title(ws, row, 'Analysis History (Last 10)', 4); row += 1
        _apply_header_row(ws, row, ['Section', 'Risk Score', 'Risk Level', 'Analyzed On']); row += 1
        for i, r in enumerate(all_risk[:10]):
            _apply_data_row(ws, row, [
                r.section.section_type.upper() if r.section else '—',
                r.risk_score,
                r.risk_level,
                r.analyzed_at.strftime('%d %b %Y'),
            ], shade=(i % 2 == 0))
            row += 1

    # ── SHEET 9: DISCLOSURES LOG ─────────────────────────────────────────────
    def _sheet_disclosures_log(self, wb, disclosures):
        ws = wb.create_sheet('Disclosure Upload Log')
        ws.sheet_view.showGridLines = False

        row = 1
        _section_title(ws, row, 'MANDATORY DISCLOSURE UPLOAD LOG', 5); row += 2

        _apply_header_row(ws, row,
            ['S.No', 'Section', 'Academic Year', 'Status', 'Uploaded On'],
            col_widths=[8, 26, 18, 18, 28]); row += 1

        if not disclosures:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.cell(row=row, column=1, value='No disclosures uploaded yet.').font = _font(color='888888', size=9)
            return

        sections_done = set()
        for i, d in enumerate(disclosures):
            _apply_data_row(ws, row, [
                i + 1,
                d.section_type.replace('_', ' ').title(),
                d.academic_year,
                d.status,
                d.uploaded_at.strftime('%d %b %Y, %I:%M %p'),
            ], shade=(i % 2 == 0))
            ws.cell(row=row, column=4).font = _font(
                bold=True, color=GREEN_HEX if d.status == 'Analyzed' else AMBER_HEX, size=9)
            sections_done.add(d.section_type)
            row += 1

        # Show missing sections
        all_sections = {'faculty', 'labs', 'infrastructure', 'students', 'financials', 'accreditation'}
        missing = all_sections - sections_done
        if missing:
            row += 1
            _section_title(ws, row, 'PENDING SECTIONS (NOT YET UPLOADED)', 5); row += 1
            _apply_header_row(ws, row, ['S.No', 'Section', 'Status', '', '']); row += 1
            for i, sec in enumerate(sorted(missing)):
                _apply_data_row(ws, row, [i + 1, sec.replace('_', ' ').title(), 'Not Uploaded', '', ''], shade=True)
                ws.cell(row=row, column=3).font = _font(bold=True, color=RED_HEX, size=9)
                row += 1


import json
import fitz  # PyMuPDF
from groq import Groq
from django.conf import settings
from django.http import JsonResponse
from django.views import View
from django.utils import timezone


VALID_SECTIONS = {'faculty', 'labs', 'infrastructure', 'students', 'financials', 'accreditation'}


# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _get_institution(request):
    inst_id = (
        request.POST.get('institution_id') or
        request.GET.get('institution_id') or
        (json.loads(request.body).get('institution_id') if request.body else None)
    )
    if not inst_id:
        return None, JsonResponse({'error': 'institution_id required'}, status=400)
    try:
        return Institution.objects.get(pk=inst_id), None
    except Institution.DoesNotExist:
        return None, JsonResponse({'error': 'Institution not found'}, status=404)


def _extract_pdf_text(pdf_file):
    try:
        doc  = fitz.open(stream=pdf_file.read(), filetype='pdf')
        text = ''.join(page.get_text() for page in doc)
        doc.close()
        return text.strip()
    except Exception:
        return ''


def _build_ai_prompt(section_type, text):
    base = (
        "You are an AICTE compliance data extractor. "
        "Extract structured data from this mandatory disclosure PDF text. "
        "Return ONLY valid JSON — no markdown, no explanation.\n\n"
    )
    prompts = {
        'faculty': base + '''Extract faculty details. JSON schema:
{"total_faculty":0,"faculty_phd_count":0,"required_faculty":0,
 "faculty_details":[{"name":"","dept":"","qualification":"","experience_years":0}]}
PDF TEXT:\n''' + text[:4000],

        'labs': base + '''Extract laboratory details. JSON schema:
{"total_labs":0,"lab_details":[{"name":"","dept":"","area_sqft":0,"equipment_count":0}]}
PDF TEXT:\n''' + text[:4000],

        'infrastructure': base + '''Extract infrastructure details. JSON schema:
{"total_classrooms":0,"library_books":0,"computer_count":0,"total_area_sqft":0,"hostel_capacity":0}
PDF TEXT:\n''' + text[:4000],

        'students': base + '''Extract student enrollment details. JSON schema:
{"total_students":0,"ug_students":0,"pg_students":0,"programs_offered":["program name"]}
PDF TEXT:\n''' + text[:4000],

        'financials': base + '''Extract financial details. JSON schema:
{"annual_budget":0.0,"fee_structure":{"ug_fee":0,"pg_fee":0}}
PDF TEXT:\n''' + text[:4000],

        'accreditation': base + '''Extract accreditation details. JSON schema:
{"naac_grade":"","nba_programs":"","iso_certified":false}
PDF TEXT:\n''' + text[:4000],
    }
    return prompts.get(section_type, base + f'Extract relevant data as JSON.\nPDF TEXT:\n' + text[:4000])


def _call_groq(prompt):
    try:
        resp = groq_client.chat.completions.create(
            model='llama-3.1-8b-instant',
            messages=[{'role': 'user', 'content': prompt}],
            max_tokens=1200,
            temperature=0.1,
        )
        raw = resp.choices[0].message.content.strip()
        raw = raw.replace('```json', '').replace('```', '').strip()
        return json.loads(raw)
    except Exception:
        return {}


def _default_section_data(section_type):
    defaults = {
        'faculty':        {'total_faculty': 0, 'faculty_phd_count': 0, 'required_faculty': 0, 'faculty_details': []},
        'labs':           {'total_labs': 0, 'lab_details': []},
        'infrastructure': {'total_classrooms': 0, 'library_books': 0, 'computer_count': 0, 'total_area_sqft': 0, 'hostel_capacity': 0},
        'students':       {'total_students': 0, 'ug_students': 0, 'pg_students': 0, 'programs_offered': []},
        'financials':     {'annual_budget': 0.0, 'fee_structure': {}},
        'accreditation':  {'naac_grade': '', 'nba_programs': '', 'iso_certified': False},
    }
    return defaults.get(section_type, {})


def _validate_section_match(section_type, ai_data):
    """
    Check whether AI-extracted data actually matches the claimed section type.
    Returns (is_valid, warning_message).
    """
    if not ai_data:
        return True, None   # empty data is OK — defaults applied

    key_map = {
        'faculty':        ['total_faculty', 'faculty_details', 'faculty_phd_count'],
        'labs':           ['total_labs', 'lab_details'],
        'infrastructure': ['total_classrooms', 'library_books', 'computer_count', 'total_area_sqft'],
        'students':       ['total_students', 'ug_students', 'pg_students', 'programs_offered'],
        'financials':     ['annual_budget', 'fee_structure'],
        'accreditation':  ['naac_grade', 'nba_programs', 'iso_certified'],
    }
    expected_keys = key_map.get(section_type, [])
    matched = sum(1 for k in expected_keys if k in ai_data)

    # Wrong-file detection: if no expected keys matched but other section's keys exist
    all_keys = {k for keys in key_map.values() for k in keys}
    present  = [k for k in ai_data if k in all_keys]

    if matched == 0 and len(present) > 0:
        # Find which section the data belongs to
        for sec, keys in key_map.items():
            if sec != section_type and sum(1 for k in keys if k in ai_data) >= 2:
                return False, (
                    f"Wrong file uploaded! You selected '{section_type}' section but the PDF "
                    f"appears to contain '{sec}' data. Please upload the correct PDF."
                )
    return True, None


def _merge_ai_data_to_institution(institution, section_type, ai_data):
    inst_data, _ = InstitutionData.objects.get_or_create(institution=institution)

    if section_type == 'faculty':
        if ai_data.get('total_faculty', 0) > 0:
            inst_data.total_faculty     = ai_data['total_faculty']
        if ai_data.get('required_faculty', 0) > 0:
            inst_data.required_faculty  = ai_data['required_faculty']
        if ai_data.get('faculty_phd_count', 0) > 0:
            inst_data.faculty_phd_count = ai_data['faculty_phd_count']
        if ai_data.get('faculty_details'):
            inst_data.faculty_details   = ai_data['faculty_details']

    elif section_type == 'labs':
        if ai_data.get('total_labs', 0) > 0:
            inst_data.total_labs = ai_data['total_labs']
        if ai_data.get('lab_details'):
            inst_data.lab_details = ai_data['lab_details']

    elif section_type == 'infrastructure':
        for field in ['total_classrooms', 'library_books', 'computer_count', 'total_area_sqft', 'hostel_capacity']:
            val = ai_data.get(field, 0)
            if val > 0:
                setattr(inst_data, field, val)

    elif section_type == 'students':
        for field in ['total_students', 'ug_students', 'pg_students']:
            val = ai_data.get(field, 0)
            if val > 0:
                setattr(inst_data, field, val)
        if ai_data.get('programs_offered'):
            inst_data.programs_offered = ai_data['programs_offered']

    elif section_type == 'financials':
        if ai_data.get('annual_budget', 0) > 0:
            inst_data.annual_budget = ai_data['annual_budget']
        if ai_data.get('fee_structure'):
            inst_data.fee_structure = ai_data['fee_structure']

    elif section_type == 'accreditation':
        if ai_data.get('naac_grade'):
            inst_data.naac_grade   = ai_data['naac_grade']
        if ai_data.get('nba_programs'):
            inst_data.nba_programs = ai_data['nba_programs']
        inst_data.iso_certified = ai_data.get('iso_certified', inst_data.iso_certified)

    inst_data.save()
    return inst_data


def _compute_risk(institution):
    """Full 16-check risk engine with per-section breakdown."""
    try:
        d = institution.data
    except InstitutionData.DoesNotExist:
        return None

    score        = 0
    factors      = []
    suggestions  = []
    section_scores = {}   # per-section contribution

    faculty_shortage = False
    infra_deficit    = False
    expired_certs    = False

    total_f  = d.total_faculty
    req_f    = d.required_faculty
    students = d.total_students
    ratio    = (students / total_f) if total_f > 0 else 0

    # ── FACULTY ───────────────────────────────────────────────────────────────
    fac_pts = 0
    if req_f > 0 and total_f < req_f:
        deficit = req_f - total_f
        fac_pts += 25; faculty_shortage = True
        factors.append(f"Faculty shortage: {total_f} available, {req_f} required (deficit: {deficit})")
        suggestions.append(f"Appoint {deficit} additional faculty to meet AICTE norms.")
    elif ratio > 20:
        fac_pts += 20; faculty_shortage = True
        factors.append(f"High student-faculty ratio 1:{ratio:.0f} (AICTE norm: 1:15)")
        suggestions.append("Recruit faculty to bring ratio below 1:15.")

    if total_f > 0:
        phd_pct = (d.faculty_phd_count / total_f) * 100
        if phd_pct < 30:
            needed = max(0, int(0.30 * total_f) - d.faculty_phd_count)
            fac_pts += 10
            factors.append(f"PhD faculty {phd_pct:.1f}% below 30% AICTE norm")
            suggestions.append(f"Encourage {needed} more faculty to pursue PhD.")

    # Duplicate faculty detection
    names = [f.get('name', '').strip().lower() for f in (d.faculty_details or []) if f.get('name')]
    duplicates = len(names) - len(set(names))
    if duplicates > 0:
        fac_pts += min(5, duplicates * 2)
        factors.append(f"Duplicate faculty entries detected: {duplicates} duplicate names in faculty list")
        suggestions.append("Audit faculty list and remove duplicate entries before resubmission.")

    section_scores['faculty'] = fac_pts
    score += fac_pts

    # ── INFRASTRUCTURE ────────────────────────────────────────────────────────
    infra_pts = 0
    if students > 0 and d.total_area_sqft > 0:
        area_per = d.total_area_sqft / students
        if area_per < 15:
            infra_pts += 20; infra_deficit = True
            factors.append(f"Area {area_per:.1f} sqft/student below 15 sqft AICTE norm")
            suggestions.append("Expand built-up area to meet 15 sqft/student minimum.")

    if d.total_labs == 0:
        infra_pts += 15; infra_deficit = True
        factors.append("No lab data uploaded — adequacy cannot be verified")
        suggestions.append("Upload Laboratory Details PDF.")

    if d.total_classrooms == 0 and students > 0:
        infra_pts += 10; infra_deficit = True
        factors.append("Classroom data missing")
        suggestions.append("Upload Infrastructure PDF with classroom details.")

    if students > 0 and d.computer_count > 0:
        comp_ratio = students / d.computer_count
        if comp_ratio > 3:
            needed = int(students / 3) - d.computer_count
            infra_pts += 10; infra_deficit = True
            factors.append(f"Computer ratio 1:{comp_ratio:.1f} exceeds 1:3 AICTE norm")
            suggestions.append(f"Procure {needed} additional computers.")

    if students > 0 and d.library_books < 10000:
        needed_books = 10000 - d.library_books
        infra_pts += 8
        factors.append(f"Library {d.library_books} volumes below 10,000 minimum")
        suggestions.append(f"Procure {needed_books} more books.")

    # Intake vs resource mismatch
    if students > 0 and d.total_classrooms > 0:
        students_per_class = students / d.total_classrooms
        if students_per_class > 70:
            infra_pts += 8; infra_deficit = True
            factors.append(f"Intake-Resource mismatch: {students_per_class:.0f} students/classroom (norm: ≤60)")
            suggestions.append("Increase classroom capacity or reduce intake to maintain 1:60 ratio.")

    if d.hostel_capacity == 0 and students > 500:
        infra_pts += 3
        factors.append("No hostel data for large institution")
        suggestions.append("Upload hostel details in Infrastructure PDF.")

    section_scores['infrastructure'] = infra_pts
    score += infra_pts

    # ── ACCREDITATION ─────────────────────────────────────────────────────────
    acc_pts = 0
    if not d.naac_grade:
        acc_pts += 10
        factors.append("NAAC accreditation data missing")
        suggestions.append("Upload Accreditation PDF with NAAC grade details.")

    # Expired certificate detection (AI flag or check from accreditation data)
    if expired_certs:
        acc_pts += 15
        factors.append("Critical certificates have expired or expiring within 30 days")
        suggestions.append("Renew expired certificates immediately before resubmission.")

    if not d.nba_programs:
        acc_pts += 5
        factors.append("No NBA accredited programs found")
        suggestions.append("Apply for NBA accreditation for eligible UG programs.")

    if not d.iso_certified:
        acc_pts += 3
        factors.append("ISO 9001:2015 certification not found")
        suggestions.append("Obtain ISO 9001:2015 certification.")

    section_scores['accreditation'] = acc_pts
    score += acc_pts

    # ── FINANCIALS ────────────────────────────────────────────────────────────
    fin_pts = 0
    if not d.annual_budget or d.annual_budget == 0:
        fin_pts += 8
        factors.append("Financial data not uploaded — budget unverified")
        suggestions.append("Upload Financial Details PDF.")

    section_scores['financials'] = fin_pts
    score += fin_pts

    # ── STUDENTS / COMPLETENESS ───────────────────────────────────────────────
    stu_pts = 0
    if not d.programs_offered:
        stu_pts += 5
        factors.append("No program data uploaded")
        suggestions.append("Upload Student Details PDF with program list.")

    computed = d.ug_students + d.pg_students
    if computed > 0 and abs(computed - students) > 50:
        stu_pts += 5
        factors.append(f"Student data inconsistency: UG+PG={computed} ≠ total={students}")
        suggestions.append("Resubmit student data with consistent figures.")

    section_scores['students'] = stu_pts
    score += stu_pts

    # ── HISTORICAL / APPROVAL PATTERN ────────────────────────────────────────
    prev_rejections = ApprovalRequest.objects.filter(institution=institution, status='rejected').count()
    if prev_rejections >= 2:
        pattern_pts = min(10, prev_rejections * 3)
        score += pattern_pts
        factors.append(f"Historical deficiency pattern: {prev_rejections} previous rejections on record")
        suggestions.append("Address all flagged issues systematically before next submission.")

    # Approval probability estimation
    score = min(score, 100)
    compliance_pct = max(0, 100 - score)

    sections_uploaded = DisclosureSection.objects.filter(institution=institution).values_list('section_type', flat=True).distinct().count()
    approval_probability = max(0, min(100, int(compliance_pct * 0.7 + (sections_uploaded / 6) * 30)))

    risk_level = 'High' if score >= 60 else ('Medium' if score >= 30 else 'Low')

    if not factors:
        factors     = ["All uploaded sections meet AICTE norms — institution is fully compliant."]
        suggestions = ["Continue maintaining compliance standards. Proceed with formal approval submission."]

    return {
        'risk_score':           score,
        'risk_level':           risk_level,
        'compliance_pct':       compliance_pct,
        'faculty_shortage':     faculty_shortage,
        'infra_deficit':        infra_deficit,
        'expired_certs':        expired_certs,
        'faculty_ratio':        round(ratio, 2),
        'risk_factors':         factors,
        'suggestions':          suggestions,
        'section_scores':       section_scores,
        'approval_probability': approval_probability,
    }


def _create_notification(institution, title, message, notif_type='info'):
    Notification.objects.create(
        institution=institution,
        title=title,
        message=message,
        notif_type=notif_type
    )


# ═══════════════════════════════════════════════════════════════════════════════
#  INSTITUTION VIEWS

# ═══════════════════════════════════════════════════════════════════════════════
@method_decorator(csrf_exempt, name='dispatch')
class RegisterView(View):
    def post(self, request):
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        email = data.get('email', '').strip().lower()
        if Institution.objects.filter(email=email).exists():
            return JsonResponse({'error': 'Email already registered.'}, status=400)

        inst = Institution(
            institution_name = data.get('institution_name', '').strip(),
            aicte_id         = data.get('aicte_id', ''),
            inst_type        = data.get('inst_type', 'Engineering'),
            category         = data.get('category', 'Affiliated'),
            year_established = int(data.get('year_established', 2000)),
            affiliated_univ  = data.get('affiliated_univ', ''),
            state            = data.get('state', ''),
            district         = data.get('district', ''),
            pincode          = data.get('pincode', ''),
            principal_name   = data.get('principal_name', ''),
            email            = email,
            mobile           = data.get('mobile', ''),
        )
        inst.set_password(data.get('password', ''))
        inst.save()
        InstitutionData.objects.create(institution=inst)

        _create_notification(inst,
            'Welcome to AICTE Compliance Portal',
            f'Registration successful for {inst.institution_name}. Upload all 6 mandatory disclosure PDFs to begin compliance review.',
            'info')

        return JsonResponse({'institution_id': inst.id, 'institution_name': inst.institution_name}, status=201)

@method_decorator(csrf_exempt, name='dispatch')
class LoginView(View):
    def post(self, request):
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        try:
            inst = Institution.objects.get(email=email)
        except Institution.DoesNotExist:
            return JsonResponse({'error': 'Invalid email or password.'}, status=401)

        if not inst.check_password(password):
            return JsonResponse({'error': 'Invalid email or password.'}, status=401)

        return JsonResponse({
            'institution_id':   inst.id,
            'institution_name': inst.institution_name,
            'aicte_id':         inst.aicte_id,
            'state':            inst.state,
            'approval_status':  inst.approval_status,
        })

@method_decorator(csrf_exempt, name='dispatch')
class AuthorityLoginView(View):
    def post(self, request):
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        AUTHORITY_CREDS = {
            'reviewer@aicte-india.org': 'demo1234',
            'authority@aicte.in':       'admin1234',
        }
        email    = data.get('email', '').strip().lower()
        password = data.get('password', '')
        if AUTHORITY_CREDS.get(email) == password:
            return JsonResponse({'name': 'AICTE Reviewer', 'email': email})
        return JsonResponse({'error': 'Invalid authority credentials.'}, status=401)

@method_decorator(csrf_exempt, name='dispatch')
class UploadDisclosureView(View):
    """Upload PDF for a section → AI extract → risk score → notify authority."""

    def post(self, request):
        inst, err = _get_institution(request)
        if err:
            return err

        section_type = request.POST.get('section_type', '').strip().lower()
        if section_type not in VALID_SECTIONS:
            return JsonResponse({'error': f'Invalid section_type. Must be one of: {", ".join(VALID_SECTIONS)}'}, status=400)

        pdf_file = request.FILES.get('pdf_file')
        if not pdf_file:
            return JsonResponse({'error': 'pdf_file is required.'}, status=400)

        # Validate file type by content
        if not pdf_file.name.lower().endswith('.pdf'):
            return JsonResponse({'error': 'Only PDF files are accepted. Please upload a .pdf file.'}, status=400)

        academic_year = request.POST.get('academic_year', '2024-25')

        # Extract PDF text
        extracted_text = _extract_pdf_text(pdf_file)

        # If PDF yields < 30 chars of text, use defaults
        if len(extracted_text) < 30:
            ai_data = _default_section_data(section_type)
            wrong_file_warning = None
        else:
            prompt   = _build_ai_prompt(section_type, extracted_text)
            raw_data = _call_groq(prompt)

            # Validate section match (wrong file detection)
            is_valid, wrong_file_warning = _validate_section_match(section_type, raw_data)
            if not is_valid:
                return JsonResponse({'error': wrong_file_warning}, status=422)

            # Fill missing keys with defaults
            defaults = _default_section_data(section_type)
            defaults.update({k: v for k, v in raw_data.items() if v not in (None, '', [], {})})
            ai_data = defaults

        # Save section — replace existing for same section_type
        pdf_file.seek(0)
        DisclosureSection.objects.filter(institution=inst, section_type=section_type).delete()
        disclosure = DisclosureSection.objects.create(
            institution    = inst,
            section_type   = section_type,
            academic_year  = academic_year,
            pdf_file       = pdf_file,
            extracted_text = extracted_text,
            ai_response    = ai_data,
            status         = 'Analyzed',
            review_status  = 'pending',
        )

        # Merge into InstitutionData
        _merge_ai_data_to_institution(inst, section_type, ai_data)

        # Compute fresh risk
        risk_result = _compute_risk(inst) or {}

        # Save AIRiskAnalysis
        AIRiskAnalysis.objects.create(
            institution    = inst,
            section        = disclosure,
            risk_score     = risk_result.get('risk_score', 0),
            risk_level     = risk_result.get('risk_level', 'Low'),
            compliance_pct = risk_result.get('compliance_pct', 100),
            faculty_shortage = risk_result.get('faculty_shortage', False),
            infra_deficit    = risk_result.get('infra_deficit', False),
            expired_certs    = risk_result.get('expired_certs', False),
            faculty_ratio    = risk_result.get('faculty_ratio', 0),
            risk_factors     = risk_result.get('risk_factors', []),
            suggestions      = risk_result.get('suggestions', []),
            section_scores   = risk_result.get('section_scores', {}),
        )

        # Auto-notify institution
        risk_level = risk_result.get('risk_level', 'Low')
        score      = risk_result.get('risk_score', 0)
        if risk_level in ('Medium', 'High'):
            _create_notification(inst,
                f'AI Risk Alert — {section_type.upper()} PDF',
                f'{risk_level} risk detected after uploading {section_type} PDF. Score: {score}/100. '
                f'Top issue: {(risk_result.get("risk_factors") or [""])[0]}',
                'danger' if risk_level == 'High' else 'warning')
        else:
            _create_notification(inst,
                f'{section_type.upper()} PDF Uploaded Successfully',
                f'Your {section_type} disclosure PDF has been analyzed. Risk score: {score}/100 (Low Risk).',
                'success')

        return JsonResponse({
            'disclosure_id': disclosure.id,
            'section_type':  section_type,
            'status':        'Analyzed',
            'ai_data':       ai_data,
            'risk':          risk_result,
            'wrong_file_warning': wrong_file_warning,
        }, status=201)

@method_decorator(csrf_exempt, name='dispatch')
class SubmitApprovalView(View):
    """Institution submits all uploaded sections for authority approval."""

    def post(self, request):
        inst, err = _get_institution(request)
        if err:
            return err

        sections = list(
            DisclosureSection.objects.filter(institution=inst)
            .values_list('section_type', flat=True)
            .distinct()
        )
        if not sections:
            return JsonResponse({'error': 'Upload at least one disclosure PDF before submitting for approval.'}, status=400)

        latest_risk = AIRiskAnalysis.objects.filter(institution=inst).first()

        # Create approval request
        ar = ApprovalRequest.objects.create(
            institution                = inst,
            status                     = 'submitted',
            risk_score_at_submission   = latest_risk.risk_score if latest_risk else 0,
            risk_level_at_submission   = latest_risk.risk_level if latest_risk else '',
            risk_factors_at_submission = latest_risk.risk_factors if latest_risk else [],
            sections_submitted         = sections,
        )

        # Update institution status
        inst.approval_status = 'pending'
        inst.save()

        # Notify institution
        _create_notification(inst,
            'Application Submitted for AICTE Review',
            f'Your mandatory disclosure application has been submitted to AICTE authority for review. '
            f'{len(sections)} sections submitted. Current risk score: {ar.risk_score_at_submission}/100.',
            'info')

        return JsonResponse({
            'approval_id': ar.id,
            'status':      ar.status,
            'sections':    sections,
            'message':     'Application submitted successfully. AICTE authority will review and respond.'
        }, status=201)

class ApprovalStatusView(View):
    """Institution checks its latest approval request status."""

    def get(self, request):
        inst, err = _get_institution(request)
        if err:
            return err

        requests = ApprovalRequest.objects.filter(institution=inst)
        latest   = requests.first()
        if not latest:
            return JsonResponse({'status': 'not_submitted', 'message': 'No approval request submitted yet.'})

        return JsonResponse({
            'approval_id':        latest.id,
            'status':             latest.status,
            'submitted_at':       latest.submitted_at.strftime('%d %b %Y, %I:%M %p'),
            'reviewed_at':        latest.reviewed_at.strftime('%d %b %Y, %I:%M %p') if latest.reviewed_at else None,
            'reviewed_by':        latest.reviewed_by,
            'authority_notes':    latest.authority_notes,
            'risk_at_submission': latest.risk_score_at_submission,
            'sections':           latest.sections_submitted,
            'section_decisions':  latest.section_decisions,
            'history': [{
                'id':           r.id,
                'status':       r.status,
                'submitted_at': r.submitted_at.strftime('%d %b %Y'),
                'reviewed_at':  r.reviewed_at.strftime('%d %b %Y') if r.reviewed_at else None,
                'authority_notes': r.authority_notes,
                'risk_score':   r.risk_score_at_submission,
            } for r in requests[:5]],
        })



class DashboardView(View):
    def get(self, request):
        inst, err = _get_institution(request)
        if err:
            return err

        disclosures      = DisclosureSection.objects.filter(institution=inst)
        sections_uploaded = list(disclosures.values_list('section_type', flat=True).distinct())
        latest_risk      = AIRiskAnalysis.objects.filter(institution=inst).first()
        latest_approval  = ApprovalRequest.objects.filter(institution=inst).first()

        try:
            d = inst.data
            inst_data = {
                'total_faculty':     d.total_faculty,
                'required_faculty':  d.required_faculty,
                'faculty_phd_count': d.faculty_phd_count,
                'total_labs':        d.total_labs,
                'total_classrooms':  d.total_classrooms,
                'library_books':     d.library_books,
                'computer_count':    d.computer_count,
                'total_area_sqft':   d.total_area_sqft,
                'hostel_capacity':   d.hostel_capacity,
                'total_students':    d.total_students,
                'ug_students':       d.ug_students,
                'pg_students':       d.pg_students,
                'programs_offered':  d.programs_offered,
                'naac_grade':        d.naac_grade,
                'nba_programs':      d.nba_programs,
                'iso_certified':     d.iso_certified,
                'annual_budget':     d.annual_budget,
                'fee_structure':     d.fee_structure,
                'faculty_details':   d.faculty_details,
                'lab_details':       d.lab_details,
            }
        except InstitutionData.DoesNotExist:
            inst_data = {}

        # Per-section data with review status
        section_details = {}
        for disc in disclosures:
            risk = disc.risk.first()
            section_details[disc.section_type] = {
                'uploaded_at':    disc.uploaded_at.strftime('%d %b %Y'),
                'status':         disc.status,
                'review_status':  disc.review_status,
                'review_notes':   disc.review_notes,
                'ai_data':        disc.ai_response,
                'risk_score':     risk.risk_score if risk else 0,
                'risk_level':     risk.risk_level if risk else 'N/A',
                'section_score':  (risk.section_scores or {}).get(disc.section_type, 0) if risk else 0,
            }

        return JsonResponse({
            'institution_id':    inst.id,
            'institution_name':  inst.institution_name,
            'aicte_id':          inst.aicte_id,
            'state':             inst.state,
            'inst_type':         inst.inst_type,
            'principal_name':    inst.principal_name,
            'approval_status':   inst.approval_status,
            'sections_uploaded': sections_uploaded,
            'total_uploads':     disclosures.count(),
            'section_details':   section_details,
            'inst_data':         inst_data,
            'risk_score':        latest_risk.risk_score     if latest_risk else 0,
            'risk_level':        latest_risk.risk_level     if latest_risk else 'Not Analyzed',
            'compliance_pct':    latest_risk.compliance_pct if latest_risk else 0,
            'faculty_shortage':  latest_risk.faculty_shortage if latest_risk else False,
            'infra_deficit':     latest_risk.infra_deficit    if latest_risk else False,
            'expired_certs':     latest_risk.expired_certs    if latest_risk else False,
            'faculty_ratio':     latest_risk.faculty_ratio    if latest_risk else 0,
            'risk_factors':      latest_risk.risk_factors     if latest_risk else [],
            'suggestions':       latest_risk.suggestions      if latest_risk else [],
            'section_scores':    latest_risk.section_scores   if latest_risk else {},
            'approval_probability': _compute_risk(inst).get('approval_probability', 0) if latest_risk else 0,
            'latest_approval': {
                'id':             latest_approval.id,
                'status':         latest_approval.status,
                'submitted_at':   latest_approval.submitted_at.strftime('%d %b %Y'),
                'authority_notes': latest_approval.authority_notes,
                'section_decisions': latest_approval.section_decisions,
            } if latest_approval else None,
        })


class DisclosuresListView(View):
    def get(self, request):
        inst, err = _get_institution(request)
        if err:
            return err

        discs  = DisclosureSection.objects.filter(institution=inst).order_by('-uploaded_at')
        result = []
        for d in discs:
            risk = d.risk.first()
            result.append({
                'id':            d.id,
                'section_type':  d.section_type,
                'academic_year': d.academic_year,
                'status':        d.status,
                'review_status': d.review_status,
                'review_notes':  d.review_notes,
                'uploaded_at':   d.uploaded_at.strftime('%d %b %Y, %I:%M %p'),
                'ai_data':       d.ai_response,
                'risk': {
                    'score':       risk.risk_score if risk else 0,
                    'level':       risk.risk_level if risk else 'N/A',
                    'section_score': (risk.section_scores or {}).get(d.section_type, 0) if risk else 0,
                } if risk else None,
            })
        return JsonResponse(result, safe=False)


class NotificationsView(View):
    def get(self, request):
        inst, err = _get_institution(request)
        if err:
            return err
        notifs = Notification.objects.filter(institution=inst)[:30]
        return JsonResponse([{
            'id':         n.id,
            'title':      n.title,
            'message':    n.message,
            'notif_type': n.notif_type,
            'is_read':    n.is_read,
            'created_at': n.created_at.strftime('%d %b %Y, %I:%M %p'),
        } for n in notifs], safe=False)


class AIRiskView(View):
    def get(self, request):
        inst, err = _get_institution(request)
        if err:
            return err

        analyses = AIRiskAnalysis.objects.filter(institution=inst)
        latest   = analyses.first()
        if not latest:
            return JsonResponse({'error': 'No risk analysis found. Upload disclosure PDFs first.'}, status=404)

        # Per-section breakdown
        section_breakdown = {}
        for disc in DisclosureSection.objects.filter(institution=inst):
            r = disc.risk.first()
            section_breakdown[disc.section_type] = {
                'uploaded_at':   disc.uploaded_at.strftime('%d %b %Y'),
                'review_status': disc.review_status,
                'section_score': (latest.section_scores or {}).get(disc.section_type, 0),
                'ai_data':       disc.ai_response,
            }

        # Faculty stats
        try:
            d = inst.data
            faculty_stats = {
                'total_faculty':     d.total_faculty,
                'required_faculty':  d.required_faculty,
                'shortage':          max(0, d.required_faculty - d.total_faculty),
                'phd_count':         d.faculty_phd_count,
                'phd_pct':           round((d.faculty_phd_count / d.total_faculty * 100), 1) if d.total_faculty > 0 else 0,
                'ratio':             round(d.total_students / d.total_faculty, 1) if d.total_faculty > 0 else 0,
                'faculty_details':   d.faculty_details[:10],
            }
        except InstitutionData.DoesNotExist:
            faculty_stats = {}

        history = [{
            'section_type': a.section.section_type if a.section else '—',
            'risk_score':   a.risk_score,
            'risk_level':   a.risk_level,
            'analyzed_at':  a.analyzed_at.strftime('%d %b %Y, %I:%M %p'),
        } for a in analyses[:15]]

        return JsonResponse({
            'risk_score':           latest.risk_score,
            'risk_level':           latest.risk_level,
            'compliance_pct':       latest.compliance_pct,
            'faculty_shortage':     latest.faculty_shortage,
            'infra_deficit':        latest.infra_deficit,
            'expired_certs':        latest.expired_certs,
            'faculty_ratio':        latest.faculty_ratio,
            'risk_factors':         latest.risk_factors,
            'suggestions':          latest.suggestions,
            'section_scores':       latest.section_scores,
            'section_breakdown':    section_breakdown,
            'faculty_stats':        faculty_stats,
            'approval_probability': _compute_risk(inst).get('approval_probability', 0),
            'history':              history,
        })


# ═══════════════════════════════════════════════════════════════════════════════
#  AUTHORITY VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

class AuthorityPendingApprovalsView(View):
    """List all pending approval requests for authority dashboard."""

    def get(self, request):
        status_filter = request.GET.get('status', '')
        qs = ApprovalRequest.objects.select_related('institution').all()
        if status_filter:
            qs = qs.filter(status=status_filter)

        result = []
        for ar in qs:
            inst      = ar.institution
            latest_risk = AIRiskAnalysis.objects.filter(institution=inst).first()
            try:
                d = inst.data
                stats = {
                    'total_faculty':  d.total_faculty,
                    'total_students': d.total_students,
                    'total_labs':     d.total_labs,
                    'naac_grade':     d.naac_grade,
                    'total_classrooms': d.total_classrooms,
                    'library_books':  d.library_books,
                    'faculty_details': d.faculty_details[:5],
                    'lab_details':     d.lab_details[:5],
                }
            except InstitutionData.DoesNotExist:
                stats = {}

            # Per-section data with their AI extracts
            sections_data = {}
            for disc in DisclosureSection.objects.filter(institution=inst):
                r = disc.risk.first()
                sections_data[disc.section_type] = {
                    'uploaded_at':   disc.uploaded_at.strftime('%d %b %Y'),
                    'review_status': disc.review_status,
                    'ai_data':       disc.ai_response,
                    'risk_score':    r.risk_score if r else 0,
                    'section_score': (r.section_scores or {}).get(disc.section_type, 0) if r else 0,
                }

            result.append({
                'approval_id':      ar.id,
                'institution_id':   inst.id,
                'institution_name': inst.institution_name,
                'aicte_id':         inst.aicte_id,
                'state':            inst.state,
                'inst_type':        inst.inst_type,
                'submitted_at':     ar.submitted_at.strftime('%d %b %Y, %I:%M %p'),
                'reviewed_at':      ar.reviewed_at.strftime('%d %b %Y, %I:%M %p') if ar.reviewed_at else None,
                'status':           ar.status,
                'authority_notes':  ar.authority_notes,
                'risk_score':       ar.risk_score_at_submission,
                'risk_level':       ar.risk_level_at_submission,
                'risk_factors':     ar.risk_factors_at_submission,
                'sections':         ar.sections_submitted,
                'section_decisions': ar.section_decisions,
                'sections_data':    sections_data,
                'inst_stats':       stats,
                'current_risk':     latest_risk.risk_score if latest_risk else 0,
            })
        return JsonResponse(result, safe=False)

@method_decorator(csrf_exempt, name='dispatch')
class AuthorityReviewView(View):
    """Authority approves or rejects a specific approval request."""

    def post(self, request):
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        approval_id  = data.get('approval_id')
        action       = data.get('action', '').lower()        # 'approve' or 'reject'
        notes        = data.get('notes', '').strip()
        section_decisions = data.get('section_decisions', {})  # {section: {status, notes}}

        if action not in ('approve', 'reject'):
            return JsonResponse({'error': "action must be 'approve' or 'reject'"}, status=400)

        try:
            ar = ApprovalRequest.objects.select_related('institution').get(pk=approval_id)
        except ApprovalRequest.DoesNotExist:
            return JsonResponse({'error': 'Approval request not found'}, status=404)

        inst = ar.institution
        ar.status            = 'approved' if action == 'approve' else 'rejected'
        ar.reviewed_at       = timezone.now()
        ar.reviewed_by       = 'AICTE Reviewer'
        ar.authority_notes   = notes
        ar.section_decisions = section_decisions
        ar.save()

        # Update institution approval_status
        inst.approval_status = 'approved' if action == 'approve' else 'rejected'
        inst.save()

        # Update per-section review status based on section_decisions
        for sec_type, decision in section_decisions.items():
            DisclosureSection.objects.filter(
                institution=inst, section_type=sec_type
            ).update(
                review_status=decision.get('status', 'pending'),
                review_notes=decision.get('notes', ''),
            )

        # Notify institution
        if action == 'approve':
            _create_notification(inst,
                '✅ AICTE Approval GRANTED',
                f'Congratulations! Your mandatory disclosure application has been APPROVED by AICTE. '
                f'Authority notes: {notes or "All sections meet compliance requirements."}',
                'success')
        else:
            # List rejected sections
            rejected_secs = [s for s, d in section_decisions.items() if d.get('status') == 'rejected']
            sec_list = ', '.join(rejected_secs) if rejected_secs else 'multiple sections'
            _create_notification(inst,
                '❌ AICTE Application REJECTED — Action Required',
                f'Your mandatory disclosure application has been REJECTED. '
                f'Rejected sections: {sec_list}. '
                f'Authority notes: {notes or "Please address the identified deficiencies and resubmit."}',
                'danger')

        return JsonResponse({
            'approval_id': ar.id,
            'status':      ar.status,
            'institution': inst.institution_name,
            'message':     f'Application {ar.status} successfully.',
        })


class AuthorityAllInstitutionsView(View):
    def get(self, request):
        institutions = Institution.objects.all().order_by('institution_name')
        result = []
        for inst in institutions:
            latest_risk    = AIRiskAnalysis.objects.filter(institution=inst).first()
            latest_approval = ApprovalRequest.objects.filter(institution=inst).first()
            try:
                d = inst.data
                students = d.total_students; faculty = d.total_faculty; labs = d.total_labs
            except InstitutionData.DoesNotExist:
                students = faculty = labs = 0

            result.append({
                'id':               inst.id,
                'institution_name': inst.institution_name,
                'aicte_id':         inst.aicte_id,
                'state':            inst.state,
                'inst_type':        inst.inst_type,
                'approval_status':  inst.approval_status,
                'total_students':   students,
                'total_faculty':    faculty,
                'total_labs':       labs,
                'risk_score':       latest_risk.risk_score if latest_risk else 0,
                'risk_level':       latest_risk.risk_level if latest_risk else 'Not Analyzed',
                'compliance_pct':   latest_risk.compliance_pct if latest_risk else 0,
                'risk_factors':     latest_risk.risk_factors if latest_risk else [],
                'latest_approval':  latest_approval.status if latest_approval else None,
            })

        result.sort(key=lambda x: x['risk_score'], reverse=True)
        return JsonResponse(result, safe=False)


class AuthorityStatsView(View):
    def get(self, request):
        total = Institution.objects.count()
        high = medium = low = 0
        for inst in Institution.objects.all():
            r = AIRiskAnalysis.objects.filter(institution=inst).first()
            if r:
                if r.risk_level == 'High':   high += 1
                elif r.risk_level == 'Medium': medium += 1
                else:                          low += 1

        pending  = ApprovalRequest.objects.filter(status='submitted').count()
        approved = ApprovalRequest.objects.filter(status='approved').count()
        rejected = ApprovalRequest.objects.filter(status='rejected').count()

        return JsonResponse({
            'total_institutions': total,
            'total_uploads':      DisclosureSection.objects.count(),
            'analyzed':           DisclosureSection.objects.filter(status='Analyzed').count(),
            'high_risk':          high,
            'medium_risk':        medium,
            'low_risk':           low,
            'pending_approvals':  pending,
            'approved':           approved,
            'rejected':           rejected,
        })

